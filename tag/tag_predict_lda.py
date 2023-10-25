import os
import re
import html
import unicodedata
import openpyxl
import time
import math
import jieba
import jieba.analyse
import jieba.posseg
import nltk
from collections import defaultdict
from gensim import corpora, models
from zhon.hanzi import punctuation as han_punc
from string import punctuation as ying_punc

from pallas.frameworks.content_understanding.tag.config import DictFilePath, ContentType, LanguageType


class DataProcess:
    def __init__(self):
        self.stop_cn_file = DictFilePath.stop_words_cn_path
        self.stop_en_file = DictFilePath.stop_words_en_path
        self.cn_dict_path = DictFilePath.cn_dict_path
        self.en_dict_path = DictFilePath.en_dict_path
        self.dynamic_coin_file = DictFilePath.dynamic_coin_path

        self.pos_cn = ContentType.pos_cn
        self.pos_en = ContentType.pos_en
        self.stop_words = self.load_stop()
        self.coin2class, self.own_coin = self.read_xlsx(self.dynamic_coin_file)
        self.punc = han_punc + ying_punc

        jieba.load_userdict(self.cn_dict_path)
        jieba.load_userdict(self.en_dict_path)

    def load_stop(self):
        stop_words = set()
        for line in open(self.stop_cn_file, 'r').readlines() + open(self.stop_en_file, 'r').readlines():
            line = line.strip()
            try:
                if len(line) < 2:
                    continue
                stop_words.add(line.lower())
            except:
                print('error idf_dict line:', line)
        return stop_words

    # 读取数据
    def read_xlsx(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        m = ws.max_row
        n = ws.max_column
        class_dict, coin2class = {}, {}
        own_coin = []

        for i in range(2, n + 1):
            class_dict[i] = ws.cell(row=1, column=i).value

        for i in range(2, m + 1):
            coin = ws.cell(row=i, column=1).value
            if not coin: continue
            own_coin.append(coin)
            clc_list = []
            for j in range(2, n + 1):
                match = ws.cell(row=i, column=j).value
                if not match: continue
                clc = class_dict.get(j, '')
                if clc: clc_list.append(clc)

            coin2class[coin] = clc_list
        return coin2class, own_coin

    def txt_clean(self, doc):
        doc = html.unescape(doc)
        doc = unicodedata.normalize('NFKC', doc)
        doc = re.sub(r'<[^>]+>', '', doc, 9999, re.S)
        doc = re.sub(r'<.*?>', '', doc, 9999, re.S)
        doc = re.sub(r'\r\n', '..', doc, 9999, re.S)
        doc = re.sub(r'\n', '..', doc, 9999, re.S)
        doc = re.sub(r'\t', ' ', doc, 9999, re.S)
        doc = re.sub(r'\r', ' ', doc, 9999, re.S)
        doc = re.sub(r'  ', ' ', doc, 9999, re.S)
        doc = re.sub(r'.td-.*?;}', '', doc, 9999, re.S)
        doc = re.sub(r'@media.*?}', '', doc, 9999, re.S)
        doc = re.sub(r'tdi.*?}', '', doc, 9999, re.S)
        doc = re.sub(r'-', ' ', doc, 9999, re.S)      ##
        doc = re.sub(r'  ', '', doc, 9999, re.S)
        doc = re.sub(r'; ;', '', doc, 9999, re.S)
        doc = re.sub(r'&#[0-9]*;', '', doc, 9999, re.S)
        #doc = re.sub(r'  ', '', doc, 9999, re.S)
        return doc

    def preprocess_chinese(self, text):
        """保留大小写，判断大写币种用"""
        word_cn = []
        seg = jieba.posseg.cut(text)
        for i in seg:
            # 包含标点 或者 纯数字
            if (re.search(r"[%s]+" % self.punc, i.word)) or (i.word.isdigit()):
                continue
            # 去停用词 + 词性筛选
            if (i.word not in self.stop_words) and (len(i.word) > 1) and (i.flag in self.pos_cn):
                word_cn.append(i.word)
        return word_cn

    def preprocess_english(self, text):
        seg = nltk.word_tokenize(text)
        seg_pos = nltk.pos_tag(seg)
        word_en = []
        for word_pos in seg_pos:
            # 包含标点 或者 纯数字
            if (re.search(r"[%s]+" % self.punc, word_pos[0])) or (word_pos[0].isdigit()):
                continue
            # 去停用词 + 词性筛选
            if (word_pos[0] not in self.stop_words) and (len(word_pos[0]) > 1) and (word_pos[1] in self.pos_en):
                word_en.append(word_pos[0])
        return word_en

    def preprocess_other_language(self, text):
        seg = nltk.word_tokenize(text)
        word_other = []
        for word in seg:
            # 包含标点 或者 纯数字
            if (re.search(r"[%s]+" % self.punc, word)) or (word.isdigit()):
                continue
            # 去停用词
            if (word not in self.stop_words) and (len(word) > 1):
                word_other.append(word)
        return word_other


class ContentAnalysis:
    """
    内容标签提取
    """
    def __init__(self):
        self.data_process = DataProcess()
        self.coin2class, self.own_coin = self.data_process.coin2class, self.data_process.own_coin
        self.num_topics = 200
        self.num_words = 20
        self.min_score = 0.03
        self.match_score = 0.6
        self.coin_score = 0.3
        self.tag_num = 6

        self.lda_dictionary_file = DictFilePath.lda_dictionary_path
        self.model_file = DictFilePath.lda_model_path
        self.synonym_file = DictFilePath.synonym_path
        # 可分中英文词典
        self.tag_dict_en_file = DictFilePath.tag_dict_en_path  # [cn --> cn、en]
        self.tag_dict_cn_file = DictFilePath.tag_dict_cn_path  # [en --> en]

        self.cn_pattern = re.compile(r'[\u4E00-\u9FD5]')
        self.synonym, self.tag_cn, self.tag_en, self.tags, self.coin_pattern_cn, self.coin_pattern_en = self.load_dict()
        self.dictionary = corpora.Dictionary.load(self.lda_dictionary_file)
        self.ldamodel, self.topics = self.load_model()
        self.tag_cn_pattern = re.compile('|'.join(self.tag_cn))
        self.tag_en_pattern = re.compile('|'.join([' ' + tag + ' ' for tag in sorted(list(self.tag_en), key=len, reverse=True)]))
        self.tags_pattern = re.compile('[^a-zA-Z](' + '|'.join(sorted(list(self.tags), key=len, reverse=True)) + ')[^a-zA-Z]')


    def load_dict(self):
        """加载同义词"""
        synonym = {}
        tag_cn, tag_en = set(), set()
        for line in open(self.synonym_file, 'r'):
            line = line.strip().split(';')
            if (len(line) < 2) or (len(line[0].strip()) < 1): continue
            fword = line[0].strip()
            for syn in line:
                if len(syn) < 1: continue
                syn = syn.lower()
                synonym[syn] = fword
                if not self.cn_pattern.search(syn):
                    tag_en.add(syn)
                else:
                    tag_cn.add(syn)

        """加载中英文标签集合，自有平台币种标签"""
        if os.path.isfile(self.tag_dict_cn_file):
            for line in open(self.tag_dict_cn_file, 'r'):
                line = line.strip()
                if len(line) < 1: continue
                tag_cn.add(line)
        coins = []
        if os.path.isfile(self.tag_dict_en_file):
            for line in open(self.tag_dict_en_file, 'r'):
                line = line.strip()
                if len(line) < 1: continue
                coin_judge = line.split(':')
                if len(coin_judge) == 2:
                    coins.append(line.split(':')[1])
                tag_en.add(coin_judge[-1])

        tags = set.union(tag_cn, tag_en)
        coins.sort(key=lambda x: len(x), reverse=True)

        coin_pattern_cn = re.compile('[^a-zA-Z](' + '|'.join(coins) + ')[^a-zA-Z]')
        coin_pattern_en = re.compile('|'.join([' ' + coin + ' ' for coin in coins]))
        return synonym, tag_cn, tag_en, tags, coin_pattern_cn, coin_pattern_en

    def load_model(self):
        ldamodel = ''
        topics = {}
        if os.path.isfile(self.model_file):
            ldamodel = models.LdaModel.load(self.model_file)
            topics = self.load_topic()
        return ldamodel, topics

    def load_topic(self):
        topics = {}
        topic_list = self.ldamodel.print_topics(num_topics=self.num_topics, num_words=self.num_words)
        for topic in topic_list:
            id = topic[0]
            word_list = topic[1].replace('"', '').split('+')
            word_list = [word.split('*') for word in word_list]
            topics[id] = {topic.strip(): float(score) for score, topic in word_list if float(score) > self.min_score}
        return topics

    def doc_cut(self, doc, language=0):
        """切词（保留特定词性、去停用词）"""
        doc = self.data_process.txt_clean(doc)
        language = int(language)
        if language == LanguageType.Chinese_simplify:
            tokens = self.data_process.preprocess_chinese(doc)
        elif language == LanguageType.English:
            tokens = self.data_process.preprocess_english(doc)
        else:
            tokens = self.data_process.preprocess_other_language(doc)
        return tokens

    def topic2tags(self, topicScoreList, match_dict, coin_dict, language):
        """根据主题及单词打分提取tags: lda关键词 + 标签匹配 + 币种标签"""
        tag_score = defaultdict(int)
        if language == LanguageType.Chinese_simplify:
            tag_dict = self.tags
        else:
            tag_dict = self.tag_en
        allscore = 0.0
        for topicid, score in topicScoreList:
            word_sc = self.topics.get(topicid, {})
            for word, sc in word_sc.items():
                # 同义词替换
                if word not in tag_dict:
                    continue
                word = self.synonym.get(word, word)
                tag_score[word] += score * sc
                allscore += score
        allscore = max(allscore, 0.5) * 0.8
        for word, score in match_dict.items():
            word = self.synonym.get(word, word)
            if (language != 1) and (self.cn_pattern.search(word)):
                # 非中文下的中文tag去除
                continue
            tag_score[word] += score * allscore

        for coin, score in coin_dict.items():
            # 重复币种标签过滤
            coin = self.synonym.get(coin, coin)
            if coin not in tag_score:
                tag_score[coin] = score * allscore

        tag_score = sorted(tag_score.items(), key=lambda x:x[1], reverse=True)
        tags = []
        for item in tag_score[:self.tag_num]:
            if item[1] > self.min_score:
                tags.append(item[0])
        return tags

    def tag_predict_lda(self, docs, use_model=0):
        """
        lda 模型测试
        分中英文：标签+同义词 --> 分中英文 --> 正则匹配+lda--> 同义词映射； 加权 --> 标签库过滤
        use_model=1: 使用lda模型； use_model=0：不使用模型
        """
        tag_match = []
        coin_match = []
        for doc_dict in docs:
            content_uniq_id, language, title, content = \
                doc_dict["content_id"], doc_dict["language"], doc_dict["title"], doc_dict["content"]
            doc = title + ' ' + content
            match_dict = defaultdict(int)
            coin_dict = defaultdict(int)
            if language == LanguageType.Chinese_simplify:
                match_key_list = self.tags_pattern.findall(doc)
            else:
                match_key_list = self.tag_en_pattern.findall(' ' + doc)
            if match_key_list:
                interval = 1.0 / len(match_key_list)
                for key in match_key_list:
                    match_dict[key.strip()] += interval
                tag_match.append(match_dict)
            else:
                tag_match.append(match_dict)

            if language == LanguageType.Chinese_simplify:
                match_coin_list = self.coin_pattern_cn.findall(doc)
            else:
                match_coin_list = self.coin_pattern_en.findall(' ' + doc)
            if match_coin_list:
                interval = self.coin_score / len(match_coin_list)
                for coin in match_coin_list:
                    coin_dict[coin.strip()] += interval
                coin_match.append(coin_dict)
            else:
                coin_match.append(coin_dict)

        tags_list = []
        if (not self.ldamodel) and (len(docs) < 2):
            use_model = 0

        if use_model == 1:
            doc_clean = [self.doc_cut((doc_dict["title"] + ' ' + doc_dict["content"]).lower(), language=doc_dict["language"]) for doc_dict in docs]
            corpus = [self.dictionary.doc2bow([e.lower() for e in text]) for text in doc_clean]

            """不加载LDA模型文件时使用"""
            if not self.ldamodel:
                self.num_topics = int(math.log(len(docs) + 1, 1.2))
                self.ldamodel = models.ldamodel.LdaModel(corpus, num_topics=self.num_topics, id2word=self.dictionary, passes=30)
                self.topics = self.load_topic()

            topiclists = self.ldamodel.get_document_topics(corpus, minimum_probability=1e-3)
            for topicScoreList, match_dict, coin_dict, doc_dict in zip(topiclists, tag_match, coin_match, docs):
                content_uniq_id, language = doc_dict["content_id"], doc_dict["language"]
                tags = self.topic2tags(topicScoreList, match_dict, coin_dict, language)
                tags_list.append(tags)
        else:
            for match_dict, coin_dict, doc_dict in zip(tag_match, coin_match, docs):
                topicScoreList = {}
                content_uniq_id, language = doc_dict["content_id"], doc_dict["language"]
                tags = self.topic2tags(topicScoreList, match_dict, coin_dict, language)
                tags_list.append(tags)
        return tags_list


class CoinAnalysis:
    """
    币种标签提取
    """
    def __init__(self):
        self.data_process = DataProcess()
        self.coin2class, self.own_coin = self.data_process.coin2class, self.data_process.own_coin
        self.max_num = 10
        self.synonym_list = []
        #self.coin_file = DictFilePath.own_coin_path
        self.dynamic_coin_file = DictFilePath.dynamic_coin_path
        self.synonym_file = DictFilePath.synonym_path
        self.cn_pattern = re.compile(r'[\u4E00-\u9FD5]')
        self.synonym, self.coin_pattern_cn, self.coin_pattern_en = self.load_dict()

    def load_dict(self):
        synonym = {}
        coins = self.own_coin
        for line in open(self.synonym_file, 'r'):
            line = line.strip().split(';')
            if (len(line) < 2) or (len(line[0].strip()) < 1): continue
            if line[0].strip() not in coins: continue
            fword = line[0].strip()
            for syn in line:
                if len(syn) < 1: continue
                synonym[syn] = fword
                self.synonym_list.append(syn)

        coins += self.synonym_list
        coins.sort(key=lambda x: len(x), reverse=True)
        coin_pattern_cn = re.compile('[^a-zA-Z](' + '|'.join(coins) + ')[^a-zA-Z]')
        coin_pattern_en = re.compile('|'.join([' ' + coin + ' ' for coin in coins]))
        return synonym, coin_pattern_cn, coin_pattern_en

    def doc2coins(self, doc):
        coin_tags = []
        if not self.cn_pattern.search(doc):
            match_coin_list = self.coin_pattern_en.findall(doc)
        else:
            match_coin_list = self.coin_pattern_cn.findall(doc)

        coin_num = defaultdict(int)
        if match_coin_list:
            for coin in match_coin_list:
                coin = coin.strip()
                coin = self.synonym.get(coin, coin)
                coin_num[coin] += 1
            coin_num = sorted(coin_num.items(), key=lambda x: x[1], reverse=True)
            coin_tags = [item[0] for item in coin_num[:self.max_num]]
        return coin_tags

    def coin_match(self, doc_list):
        coin_tags_list = []
        for item in doc_list:
            content_id = item.get('content_id', '')
            if not id:
                continue
            doc = ' '.join([item.get('title', ''), item.get('content', '')])
            coin_tags = self.doc2coins(doc)
            coin_tags_list.append({"content_id": content_id, "coinTags": coin_tags})
        return coin_tags_list


if __name__ == '__main__':
    """测试样例"""
    content_id = '123'
    language = 1
    title = 'Epic Games 筹集20亿美元“构建元宇宙”——它会使用加密货币或NFT吗？'
    content = '热门游戏《堡垒之夜》开发商 Epic Games 今天宣布，它已经筹集了20亿美元的资金，想要“建立元宇宙”。索尼和乐高集团的母公司KIRKBI各投资10亿美元，投后估值为 315 亿美元。'
    doc1 = {"content_id": content_id, "language": language, "title": title, "content": content}

    content_id = '456'
    language = 3
    title = 'How Bridging Crypto and Fiat Will Transform the Global Financial System: An Interview With SafeGram CEO Ivan Tomic'
    content = 'With more and more funds coming into the system, the next obvious step is to get these funds out of the system and this is where DeFi is currently lagging behind.The efforts in this direction have been minimal so far.'
    doc2 = {"content_id": content_id, "language": language, "title": title, "content": content}

    docs = [doc1, doc2]

    """内容标签测试"""
    lda = ContentAnalysis()
    tags_dict = lda.tag_predict_lda(docs, use_model=1)
    print(tags_dict)

    """币种标签测试"""
    coin_analysis = CoinAnalysis()
    coin_tags_list = coin_analysis.coin_match(docs)
    print(coin_tags_list)



